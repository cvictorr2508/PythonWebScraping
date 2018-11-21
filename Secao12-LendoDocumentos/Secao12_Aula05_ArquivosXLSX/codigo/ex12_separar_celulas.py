from openpyxl import load_workbook
from openpyxl.styles import Alignment
wb = load_workbook('MeusNumeros.xlsx')

planilha = wb['Sheet']

planilha.unmerge_cells("A2:A7")
planilha['B1'] = "NÚMEROS"
planilha['B1'].alignment = Alignment(horizontal='center')
wb.save('MeusNumeros.xlsx')