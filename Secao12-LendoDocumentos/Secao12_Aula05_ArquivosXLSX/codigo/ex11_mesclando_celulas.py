from openpyxl import load_workbook

wb = load_workbook('MeusNumeros.xlsx')

planilha = wb['Sheet']

planilha.merge_cells(start_row=1, start_column=2, end_row=1, end_column=5)
planilha.merge_cells("A2:A7")

wb.save('MeusNumeros.xlsx')