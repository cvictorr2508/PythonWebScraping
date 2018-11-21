from openpyxl import load_workbook

arquivo = 'Precos.xlsx'
wb = load_workbook(arquivo)

wb.create_sheet('Planilha de carros')

wb.save(arquivo)