from openpyxl import load_workbook

wb = load_workbook('Precos.xlsx')
planilha = wb['Tabela de preços']

a1 = planilha['A1']
print(a1.value)

a1 = planilha.cell(row=1, column=1)
print(a1.value)